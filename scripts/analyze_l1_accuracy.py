from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SOH_PATH = Path(r"C:\Users\M. Pandu Kurnia\Downloads\20260723_133134.csv")
MASTER_PATH = Path(r"C:\Users\M. Pandu Kurnia\Downloads\20260723_164604.csv")
TARGET_PATH = Path(r"C:\Users\M. Pandu Kurnia\Downloads\adadada.xlsx")
OUTPUT_PATH = Path(
    r"C:\Users\M. Pandu Kurnia\Documents\INVENTORY CONTROL TOWER\outputs\l1_accuracy_summary.json"
)


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


soh = pd.read_csv(
    SOH_PATH,
    dtype={
        "product_id": "string",
        "sku_number": "string",
        "package_label": "string",
    },
    low_memory=False,
)
master = pd.read_csv(MASTER_PATH, dtype="string")
soh["stock"] = pd.to_numeric(soh["stock"], errors="coerce").fillna(0)
soh["cogs"] = pd.to_numeric(soh["cogs"], errors="coerce").fillna(0)

canonical_categories = sorted(
    {str(value).strip() for value in soh["l1_category_name"].dropna().unique()},
    key=len,
    reverse=True,
)
canonical_by_norm = {norm(value): value for value in canonical_categories}

aliases = {
    norm("Perawatan diri"): canonical_by_norm[norm("Perawatan Diri")],
    norm("Tepung & Bahan kue"): canonical_by_norm[norm("Tepung & Bahan Kue")],
    norm("Pelaratan ibu & Bayi"): canonical_by_norm[norm("Peralatan Ibu & Bayi")],
}


def allowed_categories(raw: object) -> tuple[set[str], bool]:
    text = norm(raw)
    if not text:
        return set(), False

    is_non_halal = "non halal" in text or "non-halal" in text
    found: set[str] = set()

    for normalized, canonical in canonical_by_norm.items():
        if normalized in text:
            found.add(canonical)

    for alias, canonical in aliases.items():
        if alias in text:
            found.add(canonical)

    return found, is_non_halal


workbook = load_workbook(TARGET_PATH, data_only=True)
sheet = workbook["Sheet1"]

# key: (physical zone-floor, aisle) -> rule
rules: dict[tuple[str, int], dict] = {}

# SPR A/B/C maps directly to physical SRA1/SRB1/SRC1.
spr_zone = {"A": "SRA1", "B": "SRB1", "C": "SRC1"}
for row in range(3, 111):
    gangway = sheet.cell(row, 1).value
    aisle = sheet.cell(row, 2).value
    category_raw = sheet.cell(row, 3).value
    if not gangway or aisle is None:
        continue

    family = str(gangway).split("-", 1)[0].strip().upper()
    if family not in spr_zone:
        continue

    categories, excluded = allowed_categories(category_raw)
    key = (spr_zone[family], int(aisle))
    rules[key] = {
        "allowed": categories,
        "excluded": excluded,
        "source": f"SPR {gangway}",
        "sequence_allowed": {},
    }

# Mezzanine A-F. Module D floor 3 is physically HRA3/HRB3.
blocks = [
    ("A", 6, 7, 8, 9),
    ("B", 11, 12, 13, 14),
    ("C", 16, 17, 18, 19),
    ("D", 21, 22, 23, 24),
    ("E", 26, 27, 28, 29),
    ("F", 31, 32, 33, 34),
]

for module, level_col, aisle_col, category_col, business_zone_col in blocks:
    current_floor: int | None = None
    for row in range(3, 111):
        level_value = sheet.cell(row, level_col).value
        if level_value is not None:
            current_floor = int(level_value)

        aisle_value = sheet.cell(row, aisle_col).value
        if aisle_value is None or current_floor is None:
            continue

        category_raw = sheet.cell(row, category_col).value
        business_zone = norm(sheet.cell(row, business_zone_col).value)
        categories, excluded = allowed_categories(category_raw)

        if module == "D" and current_floor == 3 and business_zone in {"hra", "hrb"}:
            physical_zone = business_zone.upper() + "3"
        else:
            physical_zone = f"MZ{module}{current_floor}"

        key = (physical_zone, int(aisle_value))
        rules[key] = {
            "allowed": categories,
            "excluded": excluded,
            "source": f"Mezzanine {module} floor {current_floor} aisle {aisle_value}",
        "sequence_allowed": {},
        }

# Confirmed exception: Tata Rumah is allowed only on SRC1 aisle 18 sequences 13-17.
src18 = rules[("SRC1", 18)]
src18["allowed"].discard(canonical_by_norm[norm("Tata Rumah")])
for sequence in range(13, 18):
    src18["sequence_allowed"][sequence] = {canonical_by_norm[norm("Tata Rumah")]}


address_pattern = re.compile(
    r"^CBT-([A-Z]+\d+)-(\d+)-(\d+)-L(\d+)-(\d+)$"
)


def address_parts(rack_name: str) -> tuple[str, int, int, int, int] | None:
    match = address_pattern.fullmatch(str(rack_name).strip())
    if not match:
        return None
    zone, aisle, bay, level, position = match.groups()
    return zone, int(aisle), int(bay), int(level), int(position)


def target_for(rack_name: str) -> tuple[str, set[str], str]:
    parts = address_parts(rack_name)
    if parts is None:
        return "NO_TARGET", set(), ""

    zone, aisle, sequence, _, _ = parts
    rule = rules.get((zone, aisle))
    if rule is None:
        return "NO_TARGET", set(), ""
    if rule["excluded"]:
        return "EXCLUDED", set(), rule["source"]

    allowed = set(rule["allowed"])
    allowed.update(rule["sequence_allowed"].get(sequence, set()))
    if not allowed:
        return "NO_TARGET", set(), rule["source"]
    return "MAPPED", allowed, rule["source"]


evaluation_rows = []
for row in soh.itertuples(index=False):
    mapping_status, allowed, mapping_source = target_for(row.rack_name)
    actual = canonical_by_norm.get(norm(row.l1_category_name), str(row.l1_category_name).strip())

    if mapping_status == "MAPPED":
        result = "COMPLIANT" if actual in allowed else "NON_COMPLIANT"
    else:
        result = mapping_status

    evaluation_rows.append(
        {
            "rack_name": row.rack_name,
            "sku": row.sku_number,
            "actual_l1": actual,
            "allowed_l1": sorted(allowed),
            "result": result,
            "qty": float(row.stock),
            "value": float(row.stock * row.cogs),
            "source": mapping_source,
        }
    )

evaluation = pd.DataFrame(evaluation_rows)

rack_status_rows = []
by_rack = evaluation.groupby("rack_name", dropna=False)
for rack_name, group in by_rack:
    results = set(group["result"])
    if "NON_COMPLIANT" in results:
        status = "NON_COMPLIANT"
    elif results == {"COMPLIANT"}:
        status = "COMPLIANT"
    elif "COMPLIANT" in results and results <= {"COMPLIANT", "EXCLUDED"}:
        status = "COMPLIANT"
    elif "EXCLUDED" in results:
        status = "EXCLUDED"
    else:
        status = "NO_TARGET"

    rack_status_rows.append(
        {
            "rack_name": rack_name,
            "status": status,
            "qty": float(group["qty"].sum()),
            "wrong_qty": float(group.loc[group["result"] == "NON_COMPLIANT", "qty"].sum()),
            "wrong_value": float(group.loc[group["result"] == "NON_COMPLIANT", "value"].sum()),
        }
    )

rack_status = pd.DataFrame(rack_status_rows)
master_eval = master.merge(rack_status, on="rack_name", how="left")
master_eval["status"] = master_eval["status"].fillna("EMPTY")
master_eval["qty"] = master_eval["qty"].fillna(0)
master_eval["wrong_qty"] = master_eval["wrong_qty"].fillna(0)
master_eval["wrong_value"] = master_eval["wrong_value"].fillna(0)

mapped = evaluation[evaluation["result"].isin(["COMPLIANT", "NON_COMPLIANT"])]
compliant = evaluation[evaluation["result"] == "COMPLIANT"]
non_compliant = evaluation[evaluation["result"] == "NON_COMPLIANT"]

mapped_racks = master_eval[master_eval["status"].isin(["COMPLIANT", "NON_COMPLIANT"])]
compliant_racks = master_eval[master_eval["status"] == "COMPLIANT"]

zone_summary = (
    master_eval.groupby("ZONE", dropna=False)
    .agg(
        master_racks=("rack_name", "nunique"),
        occupied_racks=("status", lambda values: int((values != "EMPTY").sum())),
        mapped_racks=("status", lambda values: int(values.isin(["COMPLIANT", "NON_COMPLIANT"]).sum())),
        wrong_racks=("status", lambda values: int((values == "NON_COMPLIANT").sum())),
        empty_racks=("status", lambda values: int((values == "EMPTY").sum())),
        no_target_racks=("status", lambda values: int((values == "NO_TARGET").sum())),
        wrong_qty=("wrong_qty", "sum"),
        wrong_value=("wrong_value", "sum"),
    )
    .reset_index()
)
zone_summary["sloc_accuracy"] = (
    1 - zone_summary["wrong_racks"] / zone_summary["mapped_racks"].replace(0, pd.NA)
)

summary = {
    "quantity_accuracy": float(compliant["qty"].sum() / mapped["qty"].sum()) if len(mapped) else None,
    "value_accuracy": float(compliant["value"].sum() / mapped["value"].sum()) if len(mapped) else None,
    "sloc_accuracy": float(len(compliant_racks) / len(mapped_racks)) if len(mapped_racks) else None,
    "mapped_qty": float(mapped["qty"].sum()),
    "compliant_qty": float(compliant["qty"].sum()),
    "wrong_qty": float(non_compliant["qty"].sum()),
    "wrong_value": float(non_compliant["value"].sum()),
    "mapped_racks": int(len(mapped_racks)),
    "compliant_racks": int(len(compliant_racks)),
    "wrong_racks": int((master_eval["status"] == "NON_COMPLIANT").sum()),
    "no_target_occupied_racks": int((master_eval["status"] == "NO_TARGET").sum()),
    "excluded_occupied_racks": int((master_eval["status"] == "EXCLUDED").sum()),
    "empty_racks": int((master_eval["status"] == "EMPTY").sum()),
    "master_racks": int(len(master_eval)),
    "rule_count": len(rules),
    "zones": zone_summary.sort_values(["wrong_qty", "wrong_racks"], ascending=False)
    .head(30)
    .where(pd.notna(zone_summary), None)
    .to_dict(orient="records"),
}

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

print(json.dumps(summary, ensure_ascii=False, indent=2))
