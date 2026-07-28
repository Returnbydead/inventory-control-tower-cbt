from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SOH_PATH = Path(r"C:\Users\M. Pandu Kurnia\Downloads\20260723_133134.csv")
TARGET_PATH = Path(r"C:\Users\M. Pandu Kurnia\Downloads\adadada.xlsx")
OUTPUT_PATH = Path(
    r"C:\Users\M. Pandu Kurnia\Documents\INVENTORY CONTROL TOWER"
    r"\public\data\l1-placement-rules.json"
)


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def main() -> None:
    soh = pd.read_csv(SOH_PATH, usecols=["l1_category_name"], low_memory=False)
    categories = sorted(
        {str(value).strip() for value in soh["l1_category_name"].dropna().unique()},
        key=len,
        reverse=True,
    )
    canonical = {norm(value): value for value in categories}
    aliases = {
        norm("Perawatan diri"): canonical[norm("Perawatan Diri")],
        norm("Tepung & Bahan kue"): canonical[norm("Tepung & Bahan Kue")],
        norm("Pelaratan ibu & Bayi"): canonical[norm("Peralatan Ibu & Bayi")],
    }

    def allowed(raw: object) -> tuple[list[str], bool]:
        text = norm(raw)
        if not text:
            return [], False
        found = {
            value for key, value in canonical.items()
            if key in text
        }
        found.update(
            value for key, value in aliases.items()
            if key in text
        )
        return sorted(found), "non halal" in text or "non-halal" in text

    workbook = load_workbook(TARGET_PATH, data_only=True)
    sheet = workbook["Sheet1"]
    rules: dict[str, dict] = {}

    spr_zone = {"A": "SRA1", "B": "SRB1", "C": "SRC1"}
    for row in range(3, 111):
        gangway = sheet.cell(row, 1).value
        rack_sequence = sheet.cell(row, 2).value
        if not gangway or rack_sequence is None:
            continue
        family = str(gangway).split("-", 1)[0].strip().upper()
        if family not in spr_zone:
            continue
        categories_allowed, excluded = allowed(sheet.cell(row, 3).value)
        # The workbook Aisle maps to the first number after the zone.
        aisle = int(str(gangway).rsplit("-", 1)[1])
        key = f"{spr_zone[family]}:{aisle:02d}"
        rules[key] = {
            "allowed": categories_allowed,
            "excluded": excluded,
            "source": f"SPR {gangway}",
            "sequence_allowed": {},
        }

    blocks = [
        ("A", 6, 7, 8, 9),
        ("B", 11, 12, 13, 14),
        ("C", 16, 17, 18, 19),
        ("D", 21, 22, 23, 24),
        ("E", 26, 27, 28, 29),
        ("F", 31, 32, 33, 34),
    ]
    for module, level_col, rack_sequence_col, category_col, business_zone_col in blocks:
        current_floor = None
        for row in range(3, 111):
            level_value = sheet.cell(row, level_col).value
            if level_value is not None:
                current_floor = int(level_value)
            rack_sequence_value = sheet.cell(row, rack_sequence_col).value
            if rack_sequence_value is None or current_floor is None:
                continue

            business_zone = norm(sheet.cell(row, business_zone_col).value)
            if module == "D" and current_floor == 3 and business_zone in {"hra", "hrb"}:
                physical_zone = f"{business_zone.upper()}3"
            else:
                physical_zone = f"MZ{module}{current_floor}"

            categories_allowed, excluded = allowed(sheet.cell(row, category_col).value)
            key = f"{physical_zone}:{int(rack_sequence_value):02d}"
            rules[key] = {
                "allowed": categories_allowed,
                "excluded": excluded,
                "source": (
                    f"Mezzanine {module} floor {current_floor} "
                    f"rack sequence {int(rack_sequence_value)}"
                ),
                "sequence_allowed": {},
            }

    tata_rumah = canonical[norm("Tata Rumah")]
    src18 = rules["SRC1:18"]
    src18["allowed"] = [
        category for category in src18["allowed"]
        if category != tata_rumah
    ]
    src18["sequence_allowed"] = {
        str(sequence): [tata_rumah] for sequence in range(13, 18)
    }

    payload = {
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": TARGET_PATH.name,
        "category_normalization": {
            norm(value): value for value in categories
        } | aliases,
        "rules": dict(sorted(rules.items())),
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(json.dumps({
        "output": str(OUTPUT_PATH),
        "rule_count": len(rules),
        "category_count": len(categories),
        "bytes": OUTPUT_PATH.stat().st_size,
    }, indent=2))


if __name__ == "__main__":
    main()
